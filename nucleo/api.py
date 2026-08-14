import random
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
import io
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework.permissions import IsAuthenticated
from django.db import transaction
from django.db.models import Sum, Count, Q, DecimalField
from django.utils import timezone
from django.db.models.functions import Coalesce
from datetime import timedelta
from .models import (
    ConfiguracionGlobal, BorradorFactura,
    PresentacionProducto, Venta, Compra, CuentaPorCobrar, 
    CuentaPorPagar, InventarioAlmacen, SesionCaja, Cliente,
    MetodoPago, PagoCuentaCobrar, DetalleVenta, DetalleCompra,
    Producto, Proveedor, PagoCuentaPagar, ConceptoEgreso, DetalleEgresoInventario,
    Almacen,
    RutaMercado,
    RutaMercadoDetalle,
    RutaMercadoCredito,
    RutaMercadoPago,
    RutaMercadoGasto,
    TomaFisica, DetalleTomaFisica, AjusteInventario, DetalleAjusteInventario,
    InventarioAlmacen, Almacen, ConfiguracionGlobal, Producto
)
from .serializers import (
    PresentacionProductoSerializer, VentaSerializer, CompraSerializer, 
    SesionCajaSerializer, PagoCuentaCobrarSerializer, ClienteSerializer, 
    MetodoPagoSerializer, PagoCuentaPagarSerializer, ProveedorSerializer,
    ConceptoEgresoSerializer, EgresoCajaSerializer, EgresoInventarioSerializer,
    BorradorFacturaSerializer, AbonoMasivoSerializer,
    RutaMercadoSerializer,
    ImportarExcelRutaSerializer,
    TomaFisicaListSerializer, TomaFisicaDetalleSerializer, CrearTomaFisicaSerializer,
    ActualizarConteoSerializer, AjusteInventarioSerializer
)
from .permissions import IsCajeroOrSuperior, IsGerenteOrAdmin
from decimal import Decimal

class GestionCajaAPIView(APIView):
    permission_classes = [IsAuthenticated, IsCajeroOrSuperior]

    def get(self, request):
        sesion = SesionCaja.objects.filter(usuario=request.user, estado='ABIERTA').first()
        if sesion:
            serializer = SesionCajaSerializer(sesion)
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response({"mensaje": "No hay caja abierta."}, status=status.HTTP_404_NOT_FOUND)

    def post(self, request):
        if SesionCaja.objects.filter(usuario=request.user, estado='ABIERTA').exists():
            return Response({"error": "Ya tienes un turno de caja abierto."}, status=status.HTTP_400_BAD_REQUEST)

        serializer = SesionCajaSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(usuario=request.user, estado='ABIERTA')
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def put(self, request):
        sesion = SesionCaja.objects.filter(usuario=request.user, estado='ABIERTA').first()
        if not sesion:
            return Response({"error": "No tienes ninguna caja abierta para cerrar."}, status=status.HTTP_400_BAD_REQUEST)

        serializer = SesionCajaSerializer(sesion, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save(estado='CERRADA', fecha_cierre=timezone.now())
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class CatalogoPosAPIView(APIView):
    permission_classes = [IsAuthenticated, IsCajeroOrSuperior]

    def get(self, request):
        sesion = SesionCaja.objects.filter(usuario=request.user, estado='ABIERTA').first()

        if not sesion:
            return Response(
                {"mensaje": "No hay caja abierta válida para operar."}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        hoy = timezone.localtime().date()
        fecha_apertura_dia = timezone.localtime(sesion.fecha_apertura).date()

        if fecha_apertura_dia < hoy:
            return Response(
                {"mensaje": "Caja de un día anterior. Cierre obligatorio."}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        print(f"📦 CATÁLOGO SOLICITADO POR: {request.user.username} (Caja OK)")

        presentaciones = PresentacionProducto.objects.all()
        serializer = PresentacionProductoSerializer(presentaciones, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class ProcesarVentaAPIView(APIView):
    permission_classes = [IsAuthenticated, IsCajeroOrSuperior]

    @transaction.atomic
    def post(self, request):
        serializer = VentaSerializer(data=request.data, context={'request': request})

        if serializer.is_valid():
            try:
                venta = serializer.save()
                resultado_proceso = venta.procesar_venta()

                # Recargar cliente para obtener saldo a favor actualizado
                cliente = venta.cliente
                cliente.refresh_from_db()

                return Response({
                    "mensaje": "Venta procesada exitosamente",
                    "venta_id": venta.id,
                    "cajero": request.user.username,
                    "total_cobrado": venta.total_principal,
                    "tipo": venta.tipo,
                    "saldo_favor_usado": float(resultado_proceso['saldo_favor_usado']),
                    "sobrante_abono": float(resultado_proceso['sobrante_abono']),
                    "saldo_restante_cxc": float(resultado_proceso['saldo_restante']),
                    "estado_cxc": resultado_proceso['estado_cxc'],
                    "saldo_favor_cliente": float(cliente.saldo_a_favor),
                    "saldo_favor_a_nueva": float(resultado_proceso.get('saldo_favor_a_nueva', Decimal('0'))),
                    "abonos_cxc_viejas": [
                        {
                            "cxc_id": a['cxc_id'],
                            "venta_id": a['venta_id'],
                            "monto_aplicado": float(a['monto_aplicado']),
                            "saldo_restante": float(a['saldo_restante']),
                            "estado": a['estado'],
                            "origen": a.get('origen', 'EFECTIVO')
                        } for a in resultado_proceso.get('abonos_cxc_viejas', [])
                    ],
                    "abono_nueva_cxc": float(resultado_proceso.get('abono_nueva_cxc', 0)),
                    "deuda_total_cliente": float(resultado_proceso.get('deuda_total_cliente', Decimal('0'))),
                }, status=status.HTTP_201_CREATED)

            except ValueError as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                return Response({"error": "Error interno al procesar la transacción."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ProcesarCompraAPIView(APIView):
    permission_classes = [IsAuthenticated, IsGerenteOrAdmin]

    @transaction.atomic
    def post(self, request):
        serializer = CompraSerializer(data=request.data, context={'request': request})

        if serializer.is_valid():
            try:
                compra = serializer.save()
                compra.procesar_compra()

                return Response({
                    "mensaje": "Compra registrada y stock actualizado exitosamente",
                    "compra_id": compra.id,
                    "registrado_por": request.user.username,
                    "total_pagar": compra.total_principal
                }, status=status.HTTP_201_CREATED)

            except ValueError as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                return Response({"error": "Error interno al procesar la compra."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class RegistrarAbonoMasivoAPIView(APIView):
    permission_classes = [IsAuthenticated, IsCajeroOrSuperior]

    @transaction.atomic
    def post(self, request):
        serializer = AbonoMasivoSerializer(data=request.data, context={'request': request})
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        cliente_id = serializer.validated_data['cliente_id']
        pagos_data = serializer.validated_data['pagos']
        tasa_cambio = serializer.validated_data.get('tasa_cambio', Decimal('0.00'))
        guardar_saldo_favor = serializer.validated_data.get('guardar_saldo_favor', False)

        try:
            cliente = Cliente.objects.get(pk=cliente_id)
        except Cliente.DoesNotExist:
            return Response({"error": "Cliente no encontrado."}, status=status.HTTP_404_NOT_FOUND)

        total_abono_usd = Decimal('0.00')
        for pago in pagos_data:
            monto = Decimal(str(pago['monto_pagado']))
            metodo = MetodoPago.objects.get(pk=pago['metodo_id'])

            if metodo.moneda_referencia == 'SECUNDARIA' and tasa_cambio > 0:
                monto_usd = monto / tasa_cambio
            else:
                monto_usd = monto
            total_abono_usd += monto_usd

        facturas_pendientes = CuentaPorCobrar.objects.filter(
            cliente=cliente,
            estado__in=['PENDIENTE', 'VENCIDA']
        ).select_related('venta').order_by('venta__fecha')

        if not facturas_pendientes.exists() and total_abono_usd > 0:
            if not guardar_saldo_favor:
                return Response({
                    "error": "El cliente no tiene facturas pendientes. El monto excedería la deuda.",
                    "requiere_confirmacion_saldo": True,
                    "monto_sobrante": float(total_abono_usd)
                }, status=status.HTTP_400_BAD_REQUEST)
            else:
                cliente.saldo_a_favor += total_abono_usd
                cliente.save(update_fields=['saldo_a_favor'])
                return Response({
                    "mensaje": "Saldo a favor guardado exitosamente.",
                    "saldo_a_favor_total": float(cliente.saldo_a_favor),
                    "monto_guardado": float(total_abono_usd)
                }, status=status.HTTP_200_OK)

        abono_restante = total_abono_usd
        facturas_pagadas = []

        for cxc in facturas_pendientes:
            if abono_restante <= Decimal('0.00'):
                break

            saldo_antes = cxc.saldo_pendiente
            monto_aplicar = min(abono_restante, saldo_antes)

            PagoCuentaCobrar.objects.create(
                cuenta=cxc,
                usuario=request.user,
                monto_abono_principal=monto_aplicar,
                monto_entregado_secundaria=Decimal('0.00'),
                tasa_cambio_pago=tasa_cambio,
                referencia=f"Abono distribuido automáticamente"
            )

            cxc.saldo_pendiente -= monto_aplicar
            if cxc.saldo_pendiente <= Decimal('0.00'):
                cxc.saldo_pendiente = Decimal('0.00')
                cxc.estado = 'PAGADA'
            cxc.save()

            abono_restante -= monto_aplicar
            facturas_pagadas.append({
                "cxc_id": cxc.id,
                "venta_id": cxc.venta.id if cxc.venta else None,
                "monto_aplicado": float(monto_aplicar),
                "saldo_restante": float(cxc.saldo_pendiente),
                "estado": cxc.estado
            })

        saldo_sobrante = Decimal('0.00')
        if abono_restante > Decimal('0.00'):
            if guardar_saldo_favor:
                cliente.saldo_a_favor += abono_restante
                cliente.save(update_fields=['saldo_a_favor'])
                saldo_sobrante = abono_restante
            else:
                return Response({
                    "error": "El monto del abono supera la deuda total del cliente.",
                    "requiere_confirmacion_saldo": True,
                    "monto_sobrante": float(abono_restante),
                    "facturas_pagadas": facturas_pagadas
                }, status=status.HTTP_400_BAD_REQUEST)

        return Response({
            "mensaje": "Abono procesado exitosamente.",
            "total_abono": float(total_abono_usd),
            "facturas_afectadas": facturas_pagadas,
            "saldo_a_favor": float(cliente.saldo_a_favor),
            "saldo_sobrante_guardado": float(saldo_sobrante)
        }, status=status.HTTP_201_CREATED)


class RegistrarAbonoCxCAPIView(APIView):
    permission_classes = [IsAuthenticated, IsGerenteOrAdmin]

    @transaction.atomic
    def post(self, request):
        serializer = PagoCuentaCobrarSerializer(data=request.data, context={'request': request})

        if serializer.is_valid():
            try:
                pago = serializer.save()
                pago.procesar_pago()

                return Response({
                    "mensaje": "Abono registrado exitosamente",
                    "pago_id": pago.id,
                    "nuevo_saldo_pendiente": pago.cuenta.saldo_pendiente,
                    "estado_cuenta": pago.cuenta.estado
                }, status=status.HTTP_201_CREATED)

            except ValueError as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                return Response({"error": "Error interno del servidor."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class DashboardResumenAPIView(APIView):
    permission_classes = [IsAuthenticated, IsGerenteOrAdmin]

    def get(self, request):
        from datetime import datetime, time, timedelta
        from decimal import Decimal

        # timezone.localtime() respeta TIME_ZONE del .env automáticamente
        ahora = timezone.localtime()
        hoy = ahora.date()
        inicio_mes = hoy.replace(day=1)

        # Construimos rangos naive en Caracas y los convertimos a aware
        inicio_hoy = timezone.make_aware(datetime.combine(hoy, time.min))
        fin_hoy = timezone.make_aware(datetime.combine(hoy + timedelta(days=1), time.min))
        inicio_mes_dt = timezone.make_aware(datetime.combine(inicio_mes, time.min))

        # VENTAS POS (usamos __lt en lugar de __lte para evitar inclusión del límite)
        ventas_hoy = Venta.objects.filter(
            fecha__gte=inicio_hoy, fecha__lt=fin_hoy, estado='PROCESADA'
        ).aggregate(
            total_usd=Sum('total_principal'),
            cantidad_facturas=Count('id')
        )

        ventas_mes = Venta.objects.filter(
            fecha__gte=inicio_mes_dt, fecha__lt=fin_hoy, estado='PROCESADA'
        ).aggregate(total_usd=Sum('total_principal'))

        # RUTAS DE MERCADO
        rutas_hoy = RutaMercado.objects.filter(
            fecha__gte=inicio_hoy, fecha__lt=fin_hoy, estado='CERRADA'
        ).aggregate(total_usd=Sum('total_venta_usd'))

        rutas_mes = RutaMercado.objects.filter(
            fecha__gte=inicio_mes_dt, fecha__lt=fin_hoy, estado='CERRADA'
        ).aggregate(total_usd=Sum('total_venta_usd'))

        total_hoy_usd = (ventas_hoy['total_usd'] or Decimal('0.00')) + (rutas_hoy['total_usd'] or Decimal('0.00'))
        total_mes_usd = (ventas_mes['total_usd'] or Decimal('0.00')) + (rutas_mes['total_usd'] or Decimal('0.00'))

        # CXC / CXP / STOCK (sin cambios)
        cxc_pendientes = CuentaPorCobrar.objects.filter(
            estado__in=['PENDIENTE', 'VENCIDA']
        ).aggregate(
            total_deuda=Sum('saldo_pendiente'),
            clientes_deudores=Count('cliente', distinct=True)
        )

        cxp_pendientes = CuentaPorPagar.objects.filter(
            estado__in=['PENDIENTE', 'VENCIDA']
        ).aggregate(total_deuda=Sum('saldo_pendiente'))

        inventario_critico = InventarioAlmacen.objects.filter(
            stock_actual_unidades_base__lte=10
        ).select_related('producto', 'almacen').order_by('stock_actual_unidades_base')[:5]

        lista_alertas_stock = [
            {
                "producto": item.producto.nombre,
                "almacen": item.almacen.nombre,
                "stock_actual": item.stock_actual_unidades_base,
                "unidad": item.producto.unidad_medida.sigla
            }
            for item in inventario_critico
        ]

        return Response({
            "ventas": {
                "hoy_total": float(total_hoy_usd),
                "hoy_cantidad": ventas_hoy['cantidad_facturas'] or 0,
                "mes_total": float(total_mes_usd),
            },
            "finanzas": {
                "por_cobrar_total": cxc_pendientes['total_deuda'] or 0.00,
                "clientes_con_deuda": cxc_pendientes['clientes_deudores'] or 0,
                "por_pagar_total": cxp_pendientes['total_deuda'] or 0.00,
            },
            "alertas": {
                "inventario_bajo": lista_alertas_stock
            }
        }, status=status.HTTP_200_OK)

class DatosInicialesPOSAPIView(APIView):
    permission_classes = [IsAuthenticated, IsCajeroOrSuperior]

    def get(self, request):
        clientes = Cliente.objects.annotate(
            deuda_total=Coalesce(
                Sum('cuentaporcobrar__saldo_pendiente', filter=Q(cuentaporcobrar__estado__in=['PENDIENTE', 'VENCIDA'])),
                Decimal('0.00'),
                output_field=DecimalField()
            )
        ).order_by('nombre')
        metodos = MetodoPago.objects.filter(activo=True)

        return Response({
            "clientes": ClienteSerializer(clientes, many=True).data,
            "metodos_pago": MetodoPagoSerializer(metodos, many=True).data
        }, status=status.HTTP_200_OK)


class ClienteListCreateAPIView(APIView):
    permission_classes = [IsAuthenticated, IsCajeroOrSuperior]

    def get(self, request):
        clientes = Cliente.objects.annotate(
            deuda_total=Coalesce(
                Sum('cuentaporcobrar__saldo_pendiente', filter=Q(cuentaporcobrar__estado__in=['PENDIENTE', 'VENCIDA'])),
                Decimal('0.00'),
                output_field=DecimalField()
            )
        ).order_by('nombre')

        data = []
        for c in clientes:
            data.append({
                "id": c.id,
                "nombre": c.nombre,
                "documento": c.documento,
                "telefono": c.telefono,
                "limite_credito": float(c.limite_credito),
                "deuda_total": float(c.deuda_total),
                "saldo_a_favor": float(c.saldo_a_favor),
                "deuda_inicial": float(c.deuda_inicial)
            })
        return Response(data, status=status.HTTP_200_OK)

    def post(self, request):
        serializer = ClienteSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class ClienteDetalleHistorialAPIView(APIView):
    permission_classes = [IsAuthenticated, IsCajeroOrSuperior]

    def get(self, request, pk):
        try:
            cliente = Cliente.objects.get(pk=pk)
            ventas = Venta.objects.filter(cliente=cliente).order_by('-fecha')
            pagos = PagoCuentaCobrar.objects.filter(cuenta__cliente=cliente).order_by('-fecha')

            cxc_pendientes = CuentaPorCobrar.objects.filter(cliente=cliente, estado__in=['PENDIENTE', 'VENCIDA'])

            config = ConfiguracionGlobal.objects.first()
            tasa_actual = float(config.tasa_cambio_actual) if config else 1.00

            data_cliente = ClienteSerializer(cliente).data
            deuda_total = cxc_pendientes.aggregate(total=Sum('saldo_pendiente'))['total'] or 0.00

            # >>> NUEVO: Incluimos deuda_inicial en la respuesta <<<
            return Response({
                "cliente": data_cliente,
                "deuda_total": float(deuda_total),
                "limite_credito": float(cliente.limite_credito),
                "saldo_a_favor": float(cliente.saldo_a_favor),
                "deuda_inicial": float(cliente.deuda_inicial),
                "tasa_actual": tasa_actual,
                "facturas_pendientes": [{
                    "cxc_id": c.id,
                    "venta_id": c.venta.id if c.venta else None,
                    "tipo_origen": "VENTA" if c.venta else "DEUDA_INICIAL",
                    "saldo_pendiente": float(c.saldo_pendiente),
                    "fecha": c.venta.fecha if c.venta else c.fecha_vencimiento,
                    "monto_total": float(c.monto_total),
                    "tipo_venta": c.venta.tipo if c.venta else "INICIAL"
                } for c in cxc_pendientes],
                "ventas": [
                    {
                        "id": v.id, 
                        "fecha": v.fecha, 
                        "tipo": v.tipo,
                        "monto": float(v.total_principal), 
                        "estado": v.estado
                    } for v in ventas
                ],
                "pagos": [
                    {
                        "id": p.id, 
                        "fecha": p.fecha, 
                        "monto": float(p.monto_abono_principal), 
                        "referencia": p.referencia, 
                        "factura_id": p.cuenta.venta.id if p.cuenta.venta else None
                    } for p in pagos
                ]
            })
        except Cliente.DoesNotExist:
            return Response({"error": "Cliente no encontrado"}, status=404)

class ClienteUpdateAPIView(APIView):
    permission_classes = [IsAuthenticated, IsGerenteOrAdmin]

    def put(self, request, pk):
        try:
            cliente = Cliente.objects.get(pk=pk)
            serializer = ClienteSerializer(cliente, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Cliente.DoesNotExist:
            return Response({"error": "Cliente no encontrado"}, status=404)

class ProductoInventarioAPIView(APIView):
    permission_classes = [IsAuthenticated, IsGerenteOrAdmin]

    def get(self, request):
        productos = Producto.objects.all().select_related('categoria', 'unidad_medida')
        data = []
        for p in productos:
            stock_total = InventarioAlmacen.objects.filter(producto=p).aggregate(total=Sum('stock_actual_unidades_base'))['total'] or 0.00
            data.append({
                "id": p.id,
                "codigo_base": p.codigo_base,
                "nombre": p.nombre,
                "categoria": p.categoria.nombre if p.categoria else "S/C",
                "costo": float(p.costo_base_moneda_principal),
                "stock_total": float(stock_total),
                "stock_inicial": float(p.stock_inicial),
                "unidad": p.unidad_medida.sigla if p.unidad_medida else "und"
            })
        return Response(data, status=status.HTTP_200_OK)

class ProductoDetalleHistorialAPIView(APIView):
    permission_classes = [IsAuthenticated, IsGerenteOrAdmin]

    def get(self, request, pk):
        try:
            producto = Producto.objects.get(pk=pk)

            stock_almacenes = InventarioAlmacen.objects.filter(producto=producto).select_related('almacen')
            data_stock = [{"almacen": s.almacen.nombre, "cantidad": float(s.stock_actual_unidades_base)} for s in stock_almacenes]

            presentaciones = PresentacionProducto.objects.filter(producto=producto).select_related('unidad_medida')
            data_pres = [{
                "nombre": p.unidad_medida.nombre if p.unidad_medida else "Base", 
                "factor": float(p.factor_conversion), 
                "precio_usd": float(p.precio_venta_principal),
                "costo": float(p.costo_presentacion),
                "margen": float(p.margen_ganancia_porcentaje)
            } for p in presentaciones]

            # Historial de Movimientos
            ventas = DetalleVenta.objects.filter(
                presentacion__producto=producto, 
                venta__estado='PROCESADA'
            ).select_related('venta').order_by('-venta__fecha')[:30]

            compras = DetalleCompra.objects.filter(
                presentacion__producto=producto, 
                compra__estado='PROCESADA'
            ).select_related('compra').order_by('-compra__fecha')[:30]

            egresos = DetalleEgresoInventario.objects.filter(
                presentacion__producto=producto, 
                egreso__estado='PROCESADO'
            ).select_related('egreso').order_by('-egreso__fecha')[:30]

            movimientos = []

            # >>> NUEVO: Stock inicial como primer movimiento <<<
            if producto.stock_inicial > Decimal('0.00'):
                almacen_inicial = producto.almacen_inicial.nombre if producto.almacen_inicial else "Almacén por defecto"
                movimientos.append({
                    "fecha": producto.fecha_creacion,
                    "tipo": "ENTRADA",
                    "motivo": f"Stock Inicial - {almacen_inicial}",
                    "cantidad": float(producto.stock_inicial)
                })

            for v in ventas:
                cantidad_base = v.cantidad_presentacion * v.presentacion.factor_conversion
                movimientos.append({
                    "fecha": v.venta.fecha,
                    "tipo": "SALIDA",
                    "motivo": f"Venta #{v.venta.id}",
                    "cantidad": float(cantidad_base)
                })

            for c in compras:
                cantidad_base = c.cantidad_presentacion * c.presentacion.factor_conversion
                movimientos.append({
                    "fecha": c.compra.fecha,
                    "tipo": "ENTRADA",
                    "motivo": f"Compra #{c.compra.id} ({c.compra.proveedor.nombre})",
                    "cantidad": float(cantidad_base)
                })

            # >>> NUEVO: Movimientos de Rutas de Mercado <<<
            rutas = RutaMercadoDetalle.objects.filter(
                presentacion__producto=producto,
                ruta__estado='CERRADA'
            ).select_related('ruta', 'presentacion').order_by('-ruta__fecha')[:30]

            for rm in rutas:
                cantidad_base = rm.cantidad_vendida * rm.presentacion.factor_conversion
                if cantidad_base > 0:
                    movimientos.append({
                        "fecha": rm.ruta.fecha,
                        "tipo": "SALIDA",
                        "motivo": f"Ruta Mercado #{rm.ruta.id}",
                        "cantidad": float(cantidad_base)
                    })

            movimientos.sort(key=lambda x: x['fecha'], reverse=True)

            return Response({
                "producto": {
                    "nombre": producto.nombre,
                    "codigo": producto.codigo_base,
                    "unidad": producto.unidad_medida.sigla if producto.unidad_medida else "",
                    "stock_inicial": float(producto.stock_inicial)
                },
                "stock_por_almacen": data_stock,
                "presentaciones": data_pres,
                "movimientos": movimientos
            }, status=status.HTTP_200_OK)

        except Producto.DoesNotExist:
            return Response({"error": "Producto no encontrado"}, status=status.HTTP_404_NOT_FOUND)

class ProveedorListCreateAPIView(APIView):
    permission_classes = [IsAuthenticated, IsGerenteOrAdmin]

    def get(self, request):
        proveedores = Proveedor.objects.all().order_by('nombre')
        serializer = ProveedorSerializer(proveedores, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        serializer = ProveedorSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class ProveedorUpdateAPIView(APIView):
    permission_classes = [IsAuthenticated, IsGerenteOrAdmin]

    def put(self, request, pk):
        try:
            proveedor = Proveedor.objects.get(pk=pk)
            serializer = ProveedorSerializer(proveedor, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Proveedor.DoesNotExist:
            return Response({"error": "Proveedor no encontrado"}, status=status.HTTP_404_NOT_FOUND)

class ProveedorDetalleHistorialAPIView(APIView):
    permission_classes = [IsAuthenticated, IsGerenteOrAdmin]

    def get(self, request, pk):
        try:
            proveedor = Proveedor.objects.get(pk=pk)
            compras = Compra.objects.filter(proveedor=proveedor).order_by('-fecha')
            pagos = PagoCuentaPagar.objects.filter(cuenta__proveedor=proveedor).order_by('-fecha')
            cxp_pendientes = CuentaPorPagar.objects.filter(proveedor=proveedor, estado__in=['PENDIENTE', 'VENCIDA'])

            config = ConfiguracionGlobal.objects.first()
            tasa_actual = float(config.tasa_cambio_actual) if config else 1.00

            deuda_total = cxp_pendientes.aggregate(total=Sum('saldo_pendiente'))['total'] or 0.00

            return Response({
                "proveedor": ProveedorSerializer(proveedor).data,
                "deuda_total": float(deuda_total),
                "limite_credito": float(proveedor.limite_credito),
                "tasa_actual": tasa_actual,
                "facturas_pendientes": [{
                    "cxp_id": c.id,
                    "compra_id": c.compra.id,
                    "saldo_pendiente": float(c.saldo_pendiente)
                } for c in cxp_pendientes],
                "compras": [{"id": c.id, "fecha": c.fecha, "tipo": c.tipo, "monto": float(c.total_principal), "estado": c.estado} for c in compras],
                "pagos": [{"id": p.id, "fecha": p.fecha, "monto": float(p.monto_abono_principal), "referencia": p.referencia, "factura_id": p.cuenta.compra.id} for p in pagos]
            })
        except Proveedor.DoesNotExist:
            return Response({"error": "Proveedor no encontrado"}, status=status.HTTP_404_NOT_FOUND)

class RegistrarAbonoCxPAPIView(APIView):
    permission_classes = [IsAuthenticated, IsGerenteOrAdmin]

    @transaction.atomic
    def post(self, request):
        serializer = PagoCuentaPagarSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            try:
                pago = serializer.save()
                pago.procesar_pago()
                return Response({"mensaje": "Abono registrado exitosamente"}, status=status.HTTP_201_CREATED)
            except ValueError as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class ConceptoEgresoListAPIView(APIView):
    permission_classes = [IsAuthenticated, IsCajeroOrSuperior]

    def get(self, request):
        conceptos = ConceptoEgreso.objects.filter(activo=True)
        serializer = ConceptoEgresoSerializer(conceptos, many=True)
        return Response(serializer.data)

class RegistrarEgresoCajaAPIView(APIView):
    permission_classes = [IsAuthenticated, IsCajeroOrSuperior]

    @transaction.atomic
    def post(self, request):
        serializer = EgresoCajaSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response({"mensaje": "Egreso de efectivo registrado correctamente."}, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class RegistrarEgresoInventarioAPIView(APIView):
    permission_classes = [IsAuthenticated, IsCajeroOrSuperior]

    @transaction.atomic
    def post(self, request):
        serializer = EgresoInventarioSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            try:
                egreso = serializer.save()
                egreso.procesar_egreso()
                return Response({
                    "mensaje": "Egreso de inventario procesado con éxito.",
                    "egreso_id": egreso.id
                }, status=status.HTTP_201_CREATED)
            except ValueError as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ActualizarCostosProductosAPIView(APIView):
    permission_classes = [IsAuthenticated, IsGerenteOrAdmin]

    @transaction.atomic
    def post(self, request):
        cambios = request.data.get('cambios', [])
        for cambio in cambios:
            presentacion_id = cambio.get('presentacion_id')
            nuevo_costo_pres = Decimal(str(cambio.get('nuevo_costo')))

            try:
                presentacion = PresentacionProducto.objects.select_related('producto').get(pk=presentacion_id)
                producto = presentacion.producto

                if presentacion.factor_conversion > 0:
                    producto.costo_base_moneda_principal = nuevo_costo_pres / presentacion.factor_conversion
                    producto.save(update_fields=['costo_base_moneda_principal'])
            except PresentacionProducto.DoesNotExist:
                continue

        return Response({"mensaje": "Costos maestros actualizados correctamente."}, status=status.HTTP_200_OK)

class DetalleVentaFacturaAPIView(APIView):
    permission_classes = [IsAuthenticated, IsCajeroOrSuperior]

    def get(self, request, pk):
        try:
            venta = Venta.objects.get(pk=pk)
            detalles = venta.detalles.all().select_related('presentacion__producto', 'presentacion__unidad_medida')
            pagos = venta.pagos.all().select_related('metodo')

            productos_data = []
            for d in detalles:
                productos_data.append({
                    "producto": d.presentacion.producto.nombre,
                    "presentacion": d.presentacion.unidad_medida.nombre if d.presentacion.unidad_medida else "Unidad",
                    "cantidad": float(d.cantidad_presentacion),
                    "precio_unitario": float(d.precio_unitario_aplicado),
                    "subtotal": float(d.subtotal)
                })

            pagos_data = []
            for p in pagos:
                pagos_data.append({
                    "metodo": p.metodo.nombre,
                    "monto_pagado": float(p.monto_pagado),
                    "monto_usd": float(p.monto_equivalente_principal),
                    "referencia": p.referencia or "S/R"
                })

            return Response({
                "id": venta.id,
                "fecha": venta.fecha,
                "tipo": venta.tipo,
                "estado": venta.estado,
                "subtotal_principal": float(venta.subtotal_principal),
                "total_impuestos_principal": float(venta.total_impuestos_principal),
                "total_principal": float(venta.total_principal),
                "total_secundaria": float(venta.total_secundaria),
                "tasa_cambio": float(venta.tasa_cambio_historica),
                "productos": productos_data,
                "pagos": pagos_data
            }, status=status.HTTP_200_OK)
        except Venta.DoesNotExist:
            return Response({"error": "Factura no encontrada"}, status=status.HTTP_404_NOT_FOUND)

class TasaStatusAPIView(APIView):
    permission_classes = [IsAuthenticated, IsCajeroOrSuperior]

    def get(self, request):
        config = ConfiguracionGlobal.objects.first()
        if not config:
            return Response({"error": "No existe configuración global"}, status=status.HTTP_400_BAD_REQUEST)

        hoy = timezone.localdate()
        requiere = True
        if config.tasa_actualizada_el:
            requiere = timezone.localdate(config.tasa_actualizada_el) != hoy

        return Response({
            "tasa_cambio_actual": float(config.tasa_cambio_actual),
            "tasa_actualizada_el": config.tasa_actualizada_el.isoformat() if config.tasa_actualizada_el else None,
            "requiere_actualizacion": requiere,
            "moneda_principal": config.moneda_principal,
            "moneda_secundaria": config.moneda_secundaria
        })

class ActualizarTasaAPIView(APIView):
    permission_classes = [IsAuthenticated, IsCajeroOrSuperior]

    def put(self, request):
        config = ConfiguracionGlobal.objects.first()
        if not config:
            return Response({"error": "No existe configuración global"}, status=status.HTTP_400_BAD_REQUEST)

        nueva_tasa = request.data.get('tasa_cambio_actual')
        if nueva_tasa is None:
            return Response({"error": "Debes enviar tasa_cambio_actual"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            config.tasa_cambio_actual = Decimal(str(nueva_tasa))
            config.tasa_actualizada_el = timezone.now()
            config.save()
            return Response({
                "mensaje": "Tasa actualizada correctamente",
                "tasa_cambio_actual": float(config.tasa_cambio_actual),
                "tasa_actualizada_el": config.tasa_actualizada_el.isoformat()
            })
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

class BorradorFacturaListCreateAPIView(APIView):
    permission_classes = [IsAuthenticated, IsCajeroOrSuperior]

    def get(self, request):
        borradores = BorradorFactura.objects.all().select_related('cliente', 'cajero')
        serializer = BorradorFacturaSerializer(borradores, many=True)
        return Response(serializer.data)

    def post(self, request):
        serializer = BorradorFacturaSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(cajero=request.user)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class CargarBorradorAPIView(APIView):
    permission_classes = [IsAuthenticated, IsCajeroOrSuperior]

    def post(self, request, pk):
        try:
            borrador = BorradorFactura.objects.get(pk=pk)
            data = BorradorFacturaSerializer(borrador).data
            data['carrito_json'] = borrador.carrito_json
            borrador.delete()
            return Response(data)
        except BorradorFactura.DoesNotExist:
            return Response({"error": "Borrador no encontrado"}, status=status.HTTP_404_NOT_FOUND)

# ==============================================================================
# RUTAS DE MERCADO
# ==============================================================================

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Protection
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.db import transaction


class GenerarExcelRutaAPIView(APIView):
    """
    POST /api/v1/rutas/generar-excel/
    Genera un .xlsx listo para llevar al mercado.
    Body: {"tipo": "TODOS"|"CON_STOCK"|"MANUAL", "productos_ids": [], "incluir_precio_sugerido": true, "tasa_cambio": 742.22}
    """
    permission_classes = [IsAuthenticated, IsCajeroOrSuperior]

    def post(self, request):
        tipo = request.data.get('tipo', 'TODOS')
        incluir_precio = request.data.get('incluir_precio_sugerido', True)
        tasa = Decimal(str(request.data.get('tasa_cambio', 1)))

        # Obtener presentaciones según filtro
        if tipo == 'CON_STOCK':
            presentaciones = PresentacionProducto.objects.filter(
                producto__stock_por_almacen__stock_actual_unidades_base__gt=0
            ).distinct().select_related('producto', 'unidad_medida')
        elif tipo == 'MANUAL':
            ids = request.data.get('productos_ids', [])
            presentaciones = PresentacionProducto.objects.filter(id__in=ids).select_related('producto', 'unidad_medida')
        else:
            presentaciones = PresentacionProducto.objects.all().select_related('producto', 'unidad_medida')

        if not presentaciones.exists():
            return Response({"error": "No se encontraron productos con el filtro seleccionado."}, status=400)

        wb = Workbook()
        ws = wb.active
        ws.title = "Ruta de Mercado"

        # Hoja oculta con tasa
        ws_config = wb.create_sheet(title="Config")
        ws_config['A1'] = 'TASA_CAMBIO'
        ws_config['B1'] = float(tasa)
        ws_config.sheet_state = 'hidden'

        # Encabezados (ID oculto pero presente para importación exacta)
        headers = ['ID', 'PRODUCTO', 'SALIDA', 'ENTRADA', 'VENDIDO', 'PRECIO BS', 'PRECIO $', 'TOTAL BS', 'TOTAL $']
        ws.append(headers)

        # Estilo header
        header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Filas de productos
        for idx, pres in enumerate(presentaciones, start=2):
            unidad = pres.unidad_medida.nombre if pres.unidad_medida else "Base"
            factor = int(pres.factor_conversion) if pres.factor_conversion % 1 == 0 else float(pres.factor_conversion)
            nombre = f"{pres.producto.nombre} ({unidad} x{factor})"
            precio_bs = ''
            if incluir_precio:
                # Precio sugerido: precio_venta_principal (USD) * tasa = BS
                precio_bs = float(pres.precio_venta_principal) * float(tasa)

            ws.append([
                pres.id,                      # ID (referencia exacta, no editable)
                nombre,
                None,                         # Salida (editable)
                None,                         # Entrada (editable)
                f'=C{idx}-D{idx}',            # Vendido
                precio_bs,                    # Precio BS (editable, sugerido)
                f"=F{idx}/'Config'!B$1",         # Precio $
                f'=E{idx}*F{idx}',                # Total BS
                f'=E{idx}*G{idx}',            # Total $
            ])

        ws.column_dimensions['A'].width = 8   # ID
        ws.column_dimensions['B'].width = 45  # Producto
        for col in ['C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 16

        # Proteger fórmulas (solo Salida, Entrada, Precio BS son editables)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            row[0].protection = Protection(locked=True)   # ID
            row[1].protection = Protection(locked=True)   # Producto
            row[2].protection = Protection(locked=False)  # Salida
            row[3].protection = Protection(locked=False)  # Entrada
            row[4].protection = Protection(locked=True)   # Vendido (fórmula)
            row[5].protection = Protection(locked=False)  # Precio BS
            row[6].protection = Protection(locked=True)   # Precio $ (fórmula)
            row[7].protection = Protection(locked=True)   # Total $ (fórmula)

        ws.protection.sheet = True
##        ws.protection.password = None  # Solo protección visual, sin contraseña

        # Respuesta como descarga
        fecha_str = timezone.now().strftime('%Y%m%d_%H%M')
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=ruta_mercado_{fecha_str}.xlsx'
        wb.save(response)
        return response


class ImportarExcelRutaAPIView(APIView):
    """
    POST /api/v1/rutas/importar-excel/
    Recibe un archivo .xlsx y devuelve JSON con los detalles pre-llenados.
    Form-data: archivo=<file>, tasa_cambio=742.22
    """
    permission_classes = [IsAuthenticated, IsCajeroOrSuperior]

    def post(self, request):
        archivo = request.FILES.get('archivo')
        if not archivo:
            return Response({"error": "Debes subir un archivo .xlsx"}, status=400)

        tasa = Decimal(str(request.data.get('tasa_cambio', 1)))

        try:
            from openpyxl import load_workbook
            wb = load_workbook(filename=io.BytesIO(archivo.read()), data_only=True)
        except Exception as e:
            return Response({"error": f"No se pudo leer el archivo: {str(e)}"}, status=400)

        if 'Config' in wb.sheetnames:
            ws_config = wb['Config']
            tasa_excel = ws_config['B1'].value
            if tasa_excel:
                try:
                    tasa = Decimal(str(tasa_excel))
                except:
                    pass

        ws = wb['Ruta de Mercado'] if 'Ruta de Mercado' in wb.sheetnames else wb.active

        detalles = []
        no_encontrados = []

        def safe_decimal(val, default=Decimal('0.00')):
            if val is None or val == '':
                return default
            try:
                if isinstance(val, float):
                    val = round(val, 2)
                return Decimal(str(val))
            except Exception:
                return default

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row or not row[0]:
                    continue

                presentacion_id = None
                if row[0] is not None:
                    try:
                        presentacion_id = int(row[0])
                    except (ValueError, TypeError):
                        presentacion_id = None

                nombre_producto = str(row[1]).strip() if row[1] else ''
                salida = safe_decimal(row[2])
                entrada = safe_decimal(row[3])
                precio_bs = safe_decimal(row[5])

                presentacion = None
                if presentacion_id:
                    try:
                        presentacion = PresentacionProducto.objects.select_related('producto').get(id=presentacion_id)
                    except PresentacionProducto.DoesNotExist:
                        presentacion = None

                if not presentacion and nombre_producto:
                    presentacion = PresentacionProducto.objects.filter(
                        Q(producto__nombre__icontains=nombre_producto) |
                        Q(unidad_medida__nombre__icontains=nombre_producto)
                    ).select_related('producto').first()

                if not presentacion:
                    no_encontrados.append(nombre_producto or f"ID {presentacion_id}")
                    continue

                vendido = salida - entrada if salida > entrada else Decimal('0.00')
                precio_usd = precio_bs / tasa if tasa > 0 else Decimal('0.00')
                subtotal_usd = vendido * precio_usd

                unidad_nombre = presentacion.unidad_medida.nombre if presentacion.unidad_medida else 'Base'
                nombre_completo = f"{presentacion.producto.nombre} ({unidad_nombre} x{presentacion.factor_conversion})"

                detalles.append({
                    "presentacion_id": presentacion.id,
                    "nombre_producto": nombre_completo,
                    "cantidad_salida": float(salida),
                    "cantidad_entrada": float(entrada),
                    "cantidad_vendida": float(vendido),
                    "precio_venta_bs": float(precio_bs),
                    "precio_venta_usd": float(precio_usd),
                    "subtotal_usd": float(subtotal_usd),
                })

            except Exception as e:
                print(f"⚠️ Error procesando fila {row_idx}: {e}")
                continue

        return Response({
            "tasa_cambio": float(tasa),
            "detalles_encontrados": len(detalles),
            "detalles": detalles,
            "no_encontrados": no_encontrados
        })

class RutaMercadoListCreateAPIView(APIView):
    permission_classes = [IsAuthenticated, IsCajeroOrSuperior]

    def get(self, request):
        rutas = RutaMercado.objects.all().select_related('usuario', 'almacen').prefetch_related(
            'detalles__presentacion__producto',
            'creditos__cliente',
            'pagos__metodo',
            'gastos__concepto'
        )
        serializer = RutaMercadoSerializer(rutas, many=True)
        return Response(serializer.data)

    def post(self, request):
        serializer = RutaMercadoSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class RutaMercadoDetalleAPIView(APIView):
    permission_classes = [IsAuthenticated, IsCajeroOrSuperior]

    def get(self, request, pk):
        try:
            ruta = RutaMercado.objects.prefetch_related(
                'detalles__presentacion__producto',
                'creditos__cliente',
                'pagos__metodo',
                'gastos__concepto'
            ).get(pk=pk)
            serializer = RutaMercadoSerializer(ruta)
            return Response(serializer.data)
        except RutaMercado.DoesNotExist:
            return Response({"error": "Ruta no encontrada"}, status=404)

    def put(self, request, pk):
        try:
            ruta = RutaMercado.objects.get(pk=pk)
            if ruta.estado == 'CERRADA' and 'estado' not in request.data:
                # Si está cerrada, no permitir cambiar detalles de productos
                pass  # El serializer maneja la lógica
            serializer = RutaMercadoSerializer(ruta, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data)
            return Response(serializer.errors, status=400)
        except RutaMercado.DoesNotExist:
            return Response({"error": "Ruta no encontrada"}, status=404)

    def delete(self, request, pk):
        try:
            ruta = RutaMercado.objects.get(pk=pk)
            if ruta.estado == 'CERRADA':
                return Response({"error": "No puedes eliminar una ruta cerrada."}, status=400)
            ruta.delete()
            return Response({"mensaje": "Ruta eliminada."}, status=204)
        except RutaMercado.DoesNotExist:
            return Response({"error": "Ruta no encontrada"}, status=404)


class CerrarRutaMercadoAPIView(APIView):
    permission_classes = [IsAuthenticated, IsGerenteOrAdmin]

    def post(self, request, pk):
        try:
            ruta = RutaMercado.objects.prefetch_related('detalles', 'creditos').get(pk=pk)
            ruta.cerrar_ruta()
            return Response({
                "mensaje": "Ruta cerrada exitosamente.",
                "ruta": RutaMercadoSerializer(ruta).data
            })
        except RutaMercado.DoesNotExist:
            return Response({"error": "Ruta no encontrada"}, status=404)
        except ValueError as e:
            return Response({"error": str(e)}, status=400)
        except Exception as e:
            return Response({"error": f"Error al cerrar la ruta: {str(e)}"}, status=500)


class ReabrirRutaMercadoAPIView(APIView):
    """
    Reabre una ruta cerrada (solo ADMIN/GERENTE).
    Revierte inventario y elimina CxC generadas (si no tienen pagos).
    """
    permission_classes = [IsAuthenticated, IsGerenteOrAdmin]

    def post(self, request, pk):
        try:
            with transaction.atomic():
                ruta = RutaMercado.objects.prefetch_related('detalles', 'creditos__cuenta_cobrar').get(pk=pk)

                if ruta.estado != 'CERRADA':
                    return Response({"error": "La ruta no está cerrada."}, status=400)

                # 1. Revertir inventario (sumar de vuelta lo vendido)
                for detalle in ruta.detalles.all():
                    cantidad_base = detalle.cantidad_vendida * detalle.presentacion.factor_conversion
                    if cantidad_base > 0:
                        inventario = InventarioAlmacen.objects.get(
                            producto=detalle.presentacion.producto,
                            almacen=ruta.almacen
                        )
                        inventario.stock_actual_unidades_base += cantidad_base
                        inventario.save()

                # 2. Eliminar CxC generadas (solo si no tienen pagos)
                for credito in ruta.creditos.all():
                    if credito.cuenta_cobrar:
                        # Verificar que no tenga pagos
                        if not credito.cuenta_cobrar.pagos.exists():
                            credito.cuenta_cobrar.delete()
                            credito.cuenta_cobrar = None
                            credito.save()
                        else:
                            return Response({
                                "error": f"No se puede reabrir: el crédito de {credito.cliente.nombre} ya tiene pagos registrados."
                            }, status=400)

                # 3. Cambiar estado
                ruta.estado = 'BORRADOR'
                ruta.save()

                return Response({
                    "mensaje": "Ruta reabierta. Inventario restaurado y CxC eliminadas.",
                    "ruta": RutaMercadoSerializer(ruta).data
                })
        except RutaMercado.DoesNotExist:
            return Response({"error": "Ruta no encontrada"}, status=404)
        except Exception as e:
            return Response({"error": str(e)}, status=500)

class AlmacenListAPIView(APIView):
    permission_classes = [IsAuthenticated, IsCajeroOrSuperior]

    def get(self, request):
        almacenes = Almacen.objects.filter(activo=True)
        data = [{"id": a.id, "nombre": a.nombre} for a in almacenes]
        return Response(data)

class CatalogoRutaMercadoAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        presentaciones = PresentacionProducto.objects.all().select_related(
            'producto', 'unidad_medida'
        )
        serializer = PresentacionProductoSerializer(presentaciones, many=True)
        return Response(serializer.data)

class VentasTurnoAPIView(APIView):
    """
    GET /api/v1/pos/ventas-turno/
    Devuelve solo las facturas PROCESADAS del turno (sesión de caja ABIERTA)
    del usuario logueado. Incluye resumen de pagos para la lista.
    """
    permission_classes = [IsAuthenticated, IsCajeroOrSuperior]

    def get(self, request):
        sesion = SesionCaja.objects.filter(
            usuario=request.user, estado='ABIERTA'
        ).first()

        if not sesion:
            return Response(
                {"error": "No tienes una caja abierta."},
                status=status.HTTP_400_BAD_REQUEST
            )

        ventas = Venta.objects.filter(
            sesion_caja=sesion,
            usuario=request.user,
            estado='PROCESADA'
        ).select_related('cliente').prefetch_related(
            'pagos__metodo', 'detalles'
        ).order_by('-fecha')

        data = []
        for v in ventas:
            pagos_resumen = []
            for p in v.pagos.all():
                pagos_resumen.append({
                    "metodo": p.metodo.nombre,
                    "monto": float(p.monto_pagado),
                    "moneda": p.metodo.moneda_referencia
                })

            data.append({
                "id": v.id,
                "fecha": timezone.localtime(v.fecha).strftime("%d/%m/%Y %H:%M:%S"),
                "cliente": v.cliente.nombre,
                "tipo": v.tipo,
                "total_usd": float(v.total_principal),
                "total_bs": float(v.total_secundaria),
                "tasa": float(v.tasa_cambio_historica),
                "items_count": v.detalles.count(),
                "pagos": pagos_resumen
            })

        return Response(data, status=status.HTTP_200_OK)


# ==============================================================================
# TOMA FÍSICA DE INVENTARIO - API VIEWS
# ==============================================================================

class CrearTomaFisicaAPIView(APIView):
    """
    POST /api/v1/inventarios/tomas/
    Crea una nueva toma física según el tipo seleccionado.
    """
    permission_classes = [IsAuthenticated, IsCajeroOrSuperior]

    @transaction.atomic
    def post(self, request):
        serializer = CrearTomaFisicaSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        almacen_id = serializer.validated_data['almacen_id']
        tipo = serializer.validated_data['tipo']
        cantidad_muestra = serializer.validated_data.get('cantidad_muestra', 10)
        productos_ids = serializer.validated_data.get('productos_ids', [])
        observacion = serializer.validated_data.get('observacion', '')

        try:
            almacen = Almacen.objects.get(pk=almacen_id, activo=True)
        except Almacen.DoesNotExist:
            return Response({"error": "Almacén no encontrado o inactivo."}, status=404)

        # Crear cabecera
        toma = TomaFisica.objects.create(
            almacen=almacen,
            usuario=request.user,
            tipo=tipo,
            cantidad_muestra=cantidad_muestra if tipo == 'MUESTRA_ALEATORIA' else 0,
            observacion=observacion
        )

        # Obtener productos según tipo
        inventarios = InventarioAlmacen.objects.filter(almacen=almacen).select_related('producto')

        if tipo == 'COMPLETO':
            items = list(inventarios)
        elif tipo == 'MUESTRA_ALEATORIA':
            items = list(inventarios)
            if len(items) > cantidad_muestra:
                items = random.sample(items, cantidad_muestra)
        elif tipo == 'POR_PRODUCTO':
            if not productos_ids:
                toma.delete()
                return Response({"error": "Debes seleccionar al menos un producto."}, status=400)
            items = list(inventarios.filter(producto_id__in=productos_ids))
        elif tipo == 'EXCEL':
            # Para EXCEL, creamos líneas vacías que se llenarán después
            # Por ahora creamos para todos los productos como plantilla
            items = list(inventarios)
        else:
            items = list(inventarios)

        # Crear detalles
        for inv in items:
            costo = inv.producto.costo_base_moneda_principal or Decimal('0.0000')
            DetalleTomaFisica.objects.create(
                toma_fisica=toma,
                producto=inv.producto,
                stock_teorico=inv.stock_actual_unidades_base,
                stock_fisico=Decimal('0.0000'),
                costo_unitario_snapshot=costo
            )

        toma.calcular_totales()
        toma.save()

        return Response({
            "mensaje": f"Toma física #{toma.id} creada exitosamente.",
            "toma_id": toma.id,
            "tipo": tipo,
            "lineas_creadas": len(items),
            "toma": TomaFisicaDetalleSerializer(toma).data
        }, status=status.HTTP_201_CREATED)


class TomaFisicaListAPIView(APIView):
    """
    GET /api/v1/inventarios/tomas/
    Lista todas las tomas físicas con filtros opcionales.
    """
    permission_classes = [IsAuthenticated, IsCajeroOrSuperior]

    def get(self, request):
        queryset = TomaFisica.objects.select_related('almacen', 'usuario').prefetch_related('detalles')

        estado = request.query_params.get('estado')
        tipo = request.query_params.get('tipo')
        almacen_id = request.query_params.get('almacen_id')

        if estado:
            queryset = queryset.filter(estado=estado)
        if tipo:
            queryset = queryset.filter(tipo=tipo)
        if almacen_id:
            queryset = queryset.filter(almacen_id=almacen_id)

        serializer = TomaFisicaListSerializer(queryset.order_by('-fecha_creacion'), many=True)
        return Response(serializer.data)


class TomaFisicaDetalleAPIView(APIView):
    """
    GET /api/v1/inventarios/tomas/<pk>/
    Detalle completo de una toma física.
    """
    permission_classes = [IsAuthenticated, IsCajeroOrSuperior]

    def get(self, request, pk):
        try:
            toma = TomaFisica.objects.prefetch_related(
                'detalles__producto__unidad_medida',
                'detalles__presentacion'
            ).get(pk=pk)
            serializer = TomaFisicaDetalleSerializer(toma)
            return Response(serializer.data)
        except TomaFisica.DoesNotExist:
            return Response({"error": "Toma física no encontrada."}, status=404)


class ActualizarConteoAPIView(APIView):
    """
    PUT /api/v1/inventarios/tomas/<pk>/
    Actualiza los conteos físicos de una toma en estado BORRADOR.
    """
    permission_classes = [IsAuthenticated, IsCajeroOrSuperior]

    @transaction.atomic
    def put(self, request, pk):
        try:
            toma = TomaFisica.objects.get(pk=pk)
        except TomaFisica.DoesNotExist:
            return Response({"error": "Toma física no encontrada."}, status=404)

        if toma.estado != 'BORRADOR':
            return Response({"error": "Solo se pueden editar tomas en estado BORRADOR."}, status=400)

        serializer = ActualizarConteoSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=400)

        detalles_data = serializer.validated_data['detalles']
        actualizados = 0

        for item in detalles_data:
            detalle_id = item['detalle_id']
            stock_fisico = Decimal(str(item['stock_fisico']))
            observacion = item.get('observacion_linea', '')

            try:
                detalle = DetalleTomaFisica.objects.get(pk=detalle_id, toma_fisica=toma)
                detalle.stock_fisico = stock_fisico
                if observacion:
                    detalle.observacion_linea = observacion
                detalle.save()
                actualizados += 1
            except DetalleTomaFisica.DoesNotExist:
                continue

        toma.calcular_totales()
        toma.save()

        return Response({
            "mensaje": f"{actualizados} líneas actualizadas.",
            "toma": TomaFisicaDetalleSerializer(toma).data
        })


class ProcesarTomaAPIView(APIView):
    """
    POST /api/v1/inventarios/tomas/<pk>/procesar/
    Procesa la toma física: crea ajustes y aplica movimientos de stock.
    """
    permission_classes = [IsAuthenticated, IsGerenteOrAdmin]

    @transaction.atomic
    def post(self, request, pk):
        try:
            toma = TomaFisica.objects.prefetch_related('detalles').get(pk=pk)
            toma.procesar_toma()
            return Response({
                "mensaje": f"Toma física #{toma.id} procesada exitosamente.",
                "toma": TomaFisicaDetalleSerializer(toma).data
            })
        except TomaFisica.DoesNotExist:
            return Response({"error": "Toma física no encontrada."}, status=404)
        except ValueError as e:
            return Response({"error": str(e)}, status=400)
        except Exception as e:
            return Response({"error": f"Error al procesar: {str(e)}"}, status=500)


class AnularTomaAPIView(APIView):
    """
    POST /api/v1/inventarios/tomas/<pk>/anular/
    Anula una toma física. Si fue procesada, revierte los movimientos.
    """
    permission_classes = [IsAuthenticated, IsGerenteOrAdmin]

    @transaction.atomic
    def post(self, request, pk):
        try:
            toma = TomaFisica.objects.get(pk=pk)
            toma.anular_toma()
            return Response({
                "mensaje": f"Toma física #{toma.id} anulada exitosamente.",
                "toma": TomaFisicaDetalleSerializer(toma).data
            })
        except TomaFisica.DoesNotExist:
            return Response({"error": "Toma física no encontrada."}, status=404)
        except ValueError as e:
            return Response({"error": str(e)}, status=400)


class GenerarExcelTomaAPIView(APIView):
    """
    POST /api/v1/inventarios/tomas/<pk>/generar-excel/
    Genera un Excel con la plantilla de la toma para llenar stock_fisico.
    """
    permission_classes = [IsAuthenticated, IsCajeroOrSuperior]

    def post(self, request, pk):
        try:
            toma = TomaFisica.objects.prefetch_related('detalles__producto__unidad_medida').get(pk=pk)
        except TomaFisica.DoesNotExist:
            return Response({"error": "Toma física no encontrada."}, status=404)

        if toma.estado != 'BORRADOR':
            return Response({"error": "Solo se pueden generar Excel de tomas en BORRADOR."}, status=400)

        wb = Workbook()
        ws = wb.active
        ws.title = "Toma Física"

        # Encabezados
        headers = ['DETALLE_ID', 'ID_PRODUCTO', 'CÓDIGO', 'PRODUCTO', 'UNIDAD', 'STOCK_TEÓRICO', 'STOCK_FÍSICO', 'OBSERVACIÓN']
        ws.append(headers)

        header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Filas
        for detalle in toma.detalles.all():
            ws.append([
                detalle.id,
                detalle.producto.id,
                detalle.producto.codigo_base,
                detalle.producto.nombre,
                detalle.producto.unidad_medida.sigla if detalle.producto.unidad_medida else 'und',
                float(detalle.stock_teorico),
                None,  # STOCK_FÍSICO editable
                ''
            ])

        # Anchos
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 25

        # Proteger columnas fijas
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for idx in [0, 1, 2, 3, 4, 5]:  # A-F protegidas
                row[idx].protection = Protection(locked=True)
            row[6].protection = Protection(locked=False)  # G editable
            row[7].protection = Protection(locked=False)  # H editable

        ws.protection.sheet = True

        fecha_str = timezone.now().strftime('%Y%m%d_%H%M')
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=toma_fisica_{toma.id}_{fecha_str}.xlsx'
        wb.save(response)
        return response


class ImportarExcelTomaAPIView(APIView):
    """
    POST /api/v1/inventarios/tomas/<pk>/importar-excel/
    Recibe un archivo .xlsx y actualiza los conteos de la toma.
    Form-data: archivo=<file>
    """
    permission_classes = [IsAuthenticated, IsCajeroOrSuperior]

    @transaction.atomic
    def post(self, request, pk):
        try:
            toma = TomaFisica.objects.get(pk=pk)
        except TomaFisica.DoesNotExist:
            return Response({"error": "Toma física no encontrada."}, status=404)

        if toma.estado != 'BORRADOR':
            return Response({"error": "Solo se pueden importar Excel en tomas BORRADOR."}, status=400)

        archivo = request.FILES.get('archivo')
        if not archivo:
            return Response({"error": "Debes subir un archivo .xlsx"}, status=400)

        try:
            wb = load_workbook(filename=io.BytesIO(archivo.read()), data_only=True)
        except Exception as e:
            return Response({"error": f"No se pudo leer el archivo: {str(e)}"}, status=400)

        ws = wb['Toma Física'] if 'Toma Física' in wb.sheetnames else wb.active

        actualizados = 0
        no_encontrados = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row or not row[0]:
                    continue

                detalle_id = int(row[0])
                stock_fisico_raw = row[6]
                observacion = str(row[7] or '').strip()

                if stock_fisico_raw is None or stock_fisico_raw == '':
                    continue

                stock_fisico = Decimal(str(stock_fisico_raw))

                try:
                    detalle = DetalleTomaFisica.objects.get(pk=detalle_id, toma_fisica=toma)
                    detalle.stock_fisico = stock_fisico
                    if observacion:
                        detalle.observacion_linea = observacion
                    detalle.save()
                    actualizados += 1
                except DetalleTomaFisica.DoesNotExist:
                    no_encontrados.append(str(detalle_id))

            except Exception as e:
                print(f"⚠️ Error procesando fila {row_idx}: {e}")
                continue

        toma.calcular_totales()
        toma.save()

        return Response({
            "mensaje": f"{actualizados} conteos importados exitosamente.",
            "no_encontrados": no_encontrados,
            "toma": TomaFisicaDetalleSerializer(toma).data
        })


class AjusteInventarioListAPIView(APIView):
    """
    GET /api/v1/inventarios/ajustes/
    Lista todos los ajustes de inventario históricos.
    """
    permission_classes = [IsAuthenticated, IsCajeroOrSuperior]

    def get(self, request):
        ajustes = AjusteInventario.objects.select_related('almacen', 'usuario', 'toma_fisica').prefetch_related('detalles')
        serializer = AjusteInventarioSerializer(ajustes.order_by('-fecha'), many=True)
        return Response(serializer.data)


class AjusteInventarioDetalleAPIView(APIView):
    """
    GET /api/v1/inventarios/ajustes/<pk>/
    Detalle de un ajuste específico.
    """
    permission_classes = [IsAuthenticated, IsCajeroOrSuperior]

    def get(self, request, pk):
        try:
            ajuste = AjusteInventario.objects.prefetch_related('detalles__producto').get(pk=pk)
            serializer = AjusteInventarioSerializer(ajuste)
            return Response(serializer.data)
        except AjusteInventario.DoesNotExist:
            return Response({"error": "Ajuste no encontrado."}, status=404)


class InformeTomaAPIView(APIView):
    """
    GET /api/v1/inventarios/tomas/<pk>/informe/
    Genera el informe completo de faltantes y sobrantes.
    """
    permission_classes = [IsAuthenticated, IsCajeroOrSuperior]

    def get(self, request, pk):
        try:
            toma = TomaFisica.objects.prefetch_related('detalles__producto__unidad_medida').get(pk=pk)
        except TomaFisica.DoesNotExist:
            return Response({"error": "Toma física no encontrada."}, status=404)

        detalles = toma.detalles.all()

        faltantes = []
        sobrantes = []
        cuadrados = []

        for d in detalles:
            item = {
                "producto_id": d.producto.id,
                "codigo": d.producto.codigo_base,
                "nombre": d.producto.nombre,
                "unidad": d.producto.unidad_medida.sigla if d.producto.unidad_medida else 'und',
                "stock_teorico": float(d.stock_teorico),
                "stock_fisico": float(d.stock_fisico),
                "diferencia": float(d.diferencia),
                "costo_unitario": float(d.costo_unitario_snapshot),
                "valor_diferencia_usd": float(d.subtotal_diferencia),
                "observacion": d.observacion_linea
            }

            if d.diferencia < 0:
                faltantes.append(item)
            elif d.diferencia > 0:
                sobrantes.append(item)
            else:
                cuadrados.append(item)

        valor_faltantes = sum(d.subtotal_diferencia for d in detalles if d.diferencia < 0)
        valor_sobrantes = sum(d.subtotal_diferencia for d in detalles if d.diferencia > 0)

        return Response({
            "toma_id": toma.id,
            "almacen": toma.almacen.nombre,
            "tipo": toma.get_tipo_display(),
            "estado": toma.get_estado_display(),
            "fecha_creacion": toma.fecha_creacion,
            "fecha_cierre": toma.fecha_cierre,
            "resumen": {
                "total_lineas": len(detalles),
                "faltantes": len(faltantes),
                "sobrantes": len(sobrantes),
                "cuadrados": len(cuadrados),
                "valor_faltantes_usd": float(valor_faltantes),
                "valor_sobrantes_usd": float(valor_sobrantes),
                "valor_neto_usd": float(valor_sobrantes - valor_faltantes)
            },
            "faltantes": faltantes,
            "sobrantes": sobrantes,
            "cuadrados": cuadrados
        })
